#!/usr/bin/env python3
"""Create a standardized Evidence Scan Citation Ledger workbook.

This helper script builds the Phase 2 Citation Ledger `.xlsx` with the
standard five-tab structure used by the evidence-scan skill:

1. README
2. Claims
3. References
4. Cross-Report Inconsistencies
5. Corrections Applied

Inputs are JSON files produced during claim extraction. The script validates
required fields, writes the workbook, and applies basic readability formatting.
It creates an initial Phase 2 ledger only: it does not verify claims, apply
corrections, or populate Claude QA.

The workflow usually handles up to 5 research reports, labeled R1 through R5.
If more report labels are present, the script warns but still creates the
workbook so the user can batch or prioritize manually.

Example:
    python scripts/create_ledger.py \
      --topic "Benefits chatbot evidence scan" \
      --claims claims.json \
      --references references.json \
      --inconsistencies inconsistencies.json \
      --output "Benefits chatbot evidence scan — Citation Ledger.xlsx"

No references available:
    python scripts/create_ledger.py \
      --topic "Benefits chatbot evidence scan" \
      --claims claims.json \
      --output "Benefits chatbot evidence scan — Citation Ledger.xlsx" \
      --notes "Reports had weak or missing citations; many claims may verify as [U]."
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CLAIMS_HEADERS = [
    "Report",
    "Claim ID",
    "Section",
    "Claim",
    "Source Cited",
    "Full Citation",
    "Evidence Type",
    "Stakes",
    "Verified",
    "Verification Notes",
    "Correction",
    "Claude QA",
]

REFERENCES_HEADERS = [
    "Report",
    "Reference ID",
    "Full Citation",
    "Referenced in claims",
]

INCONSISTENCY_HEADERS = [
    "Inconsistency ID",
    "Source scope",
    "Report",
    "Claim text",
    "Affected Claim IDs",
    "Type",
    "Resolution",
]

CORRECTIONS_HEADERS = [
    "Claim ID",
    "Verified code",
    "Original text",
    "Corrected text",
    "Reason",
    "Claude QA",
]

README_HEADERS = ["Field", "Value"]
COVERAGE_HEADERS = ["Report", "Section", "Pages", "Claims extracted", "Stakes breakdown"]

REQUIRED_CLAIM_FIELDS = {"Report", "Claim ID", "Claim", "Evidence Type", "Stakes"}
VALID_STAKES = {"H", "M", "L"}
INITIAL_LEDGER_BLANK_FIELDS = {"Verified", "Verification Notes", "Correction", "Claude QA"}
RECOMMENDED_REPORT_LABELS = {"R1", "R2", "R3", "R4", "R5"}

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="EAF3F8")


def load_json(path: Path | None) -> list[dict[str, Any]]:
    """Load a JSON file containing a list of objects.

    Missing optional files return an empty list.
    """
    if path is None or not path.exists():
        return []

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"{path} is not valid JSON: {exc}") from exc

    if not isinstance(data, list):
        raise ValueError(f"{path} must contain a list of objects")

    for index, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"{path} item {index} must be an object")

    return data


def normalize_row(row: dict[str, Any], headers: list[str]) -> list[str]:
    """Return row values in the workbook header order."""
    values: list[str] = []
    for header in headers:
        value = row.get(header, "")
        if isinstance(value, list):
            value = ", ".join(str(item) for item in value)
        elif value is None:
            value = ""
        values.append(str(value))
    return values


def prepare_initial_claims(claims: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return claim rows prepared for an initial Phase 2 ledger.

    Initial ledgers are verification-ready, not verified. Clear verification
    fields even if upstream extraction data accidentally included values.
    """
    prepared: list[dict[str, Any]] = []
    for claim in claims:
        row = dict(claim)
        for field in INITIAL_LEDGER_BLANK_FIELDS:
            row[field] = ""
        prepared.append(row)
    return prepared


def warn_if_more_than_five_reports(claims: list[dict[str, Any]]) -> None:
    """Warn when a ledger includes more than the usual five report labels."""
    reports = {str(claim.get("Report", "")).strip() for claim in claims if claim.get("Report")}
    recommended = reports & RECOMMENDED_REPORT_LABELS
    extra = sorted(reports - RECOMMENDED_REPORT_LABELS)
    if len(reports) > 5 or extra:
        print(
            "Warning: this ledger includes report labels beyond the usual R1 through R5 workflow. "
            "Consider batching reports or prioritizing the highest-value five before verification. "
            f"Reports found: {', '.join(sorted(reports))}",
            file=sys.stderr,
        )
    elif len(recommended) > 5:
        print(
            "Warning: this ledger includes more than five reports. Consider batching before verification.",
            file=sys.stderr,
        )


def validate_claims(claims: list[dict[str, Any]]) -> None:
    """Validate required claim fields and row-level consistency."""
    seen_ids: set[str] = set()

    for index, claim in enumerate(claims, start=1):
        missing = [field for field in REQUIRED_CLAIM_FIELDS if not str(claim.get(field, "")).strip()]
        if missing:
            raise ValueError(f"Claim row {index} missing required fields: {', '.join(missing)}")

        claim_id = str(claim.get("Claim ID", "")).strip()
        if claim_id in seen_ids:
            raise ValueError(f"Duplicate Claim ID: {claim_id}")
        seen_ids.add(claim_id)

        stakes = str(claim.get("Stakes", "")).strip()
        if stakes not in VALID_STAKES:
            raise ValueError(f"Claim row {index} has invalid Stakes value: {stakes}. Use H, M, or L.")


def validate_references(references: list[dict[str, Any]]) -> None:
    """Validate reference rows enough to prevent unusable workbooks.

    Empty reference lists are allowed because some reports have weak or missing
    citations. Those limitations should be documented in README notes and in
    the claim rows themselves.
    """
    seen_ids: set[tuple[str, str]] = set()

    for index, reference in enumerate(references, start=1):
        report = str(reference.get("Report", "")).strip()
        reference_id = str(reference.get("Reference ID", "")).strip()
        citation = str(reference.get("Full Citation", "")).strip()

        if not report or not reference_id or not citation:
            raise ValueError(
                f"Reference row {index} must include Report, Reference ID, and Full Citation"
            )

        key = (report, reference_id)
        if key in seen_ids:
            raise ValueError(f"Duplicate reference ID within report: {report} / {reference_id}")
        seen_ids.add(key)


def style_sheet(ws: Worksheet) -> None:
    """Apply basic readable formatting to a worksheet."""
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in list(column_cells)[:80]:
            max_len = max(max_len, len(str(cell.value or "")))
        width = min(max(max_len + 2, 12), 70)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def add_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[dict[str, Any]] | None = None,
) -> Worksheet:
    """Add a worksheet with headers and optional rows."""
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows or []:
        ws.append(normalize_row(row, headers))
    style_sheet(ws)
    return ws


def build_coverage_rows(claims: list[dict[str, Any]]) -> list[list[str | int]]:
    """Summarize claims by report and section for README Coverage QA."""
    summary: dict[tuple[str, str], dict[str, int]] = defaultdict(
        lambda: {"total": 0, "H": 0, "M": 0, "L": 0}
    )

    for claim in claims:
        report = str(claim.get("Report", "Unknown") or "Unknown")
        section = str(claim.get("Section", "All") or "All")
        stakes = str(claim.get("Stakes", "")).strip()
        key = (report, section)
        summary[key]["total"] += 1
        if stakes in VALID_STAKES:
            summary[key][stakes] += 1

    rows: list[list[str | int]] = []
    for (report, section), counts in sorted(summary.items()):
        rows.append(
            [
                report,
                section,
                "",
                counts["total"],
                f"{counts['H']}H / {counts['M']}M / {counts['L']}L",
            ]
        )
    return rows


def build_readme(
    wb: Workbook,
    topic: str,
    verification_tool: str,
    claims: list[dict[str, Any]],
    references: list[dict[str, Any]],
    notes: str = "",
) -> Worksheet:
    """Build the README sheet with scope metadata and Coverage QA."""
    ws = wb.active
    ws.title = "README"

    source_note = "References included" if references else "No references file or no references extracted"
    if not references:
        source_note += "; verification may return [U] for uncited claims"

    metadata_rows = [
        ["Topic", topic],
        ["Verification tool planned", verification_tool],
        ["Verification codes", "[C] Confirmed; [P] Partially confirmed; [U] Unconfirmed; [X] Corrected"],
        ["Stakes key", "H = recommendation-changing; M = supports argument; L = contextual"],
        ["Report limit note", "Usual workflow supports up to 5 reports labeled R1 through R5; batch larger sets."],
        ["Source note", source_note],
        ["Ledger stage", "Initial Phase 2 ledger; verification fields intentionally blank"],
        ["Notes", notes],
    ]

    ws.append(README_HEADERS)
    for row in metadata_rows:
        ws.append(row)

    ws.append([])
    coverage_start_row = ws.max_row + 1
    ws.append(COVERAGE_HEADERS)
    for row in build_coverage_rows(claims):
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for cell in ws[coverage_start_row]:
        cell.font = Font(bold=True)
        cell.fill = SECTION_FILL

    style_sheet(ws)
    return ws


def create_workbook(
    topic: str,
    claims: list[dict[str, Any]],
    references: list[dict[str, Any]],
    inconsistencies: list[dict[str, Any]],
    verification_tool: str,
    notes: str,
) -> Workbook:
    """Create the complete initial Citation Ledger workbook."""
    validate_claims(claims)
    validate_references(references)
    warn_if_more_than_five_reports(claims)

    initial_claims = prepare_initial_claims(claims)

    wb = Workbook()
    build_readme(wb, topic, verification_tool, initial_claims, references, notes)
    add_sheet(wb, "Claims", CLAIMS_HEADERS, initial_claims)
    add_sheet(wb, "References", REFERENCES_HEADERS, references)
    add_sheet(wb, "Cross-Report Inconsistencies", INCONSISTENCY_HEADERS, inconsistencies)
    add_sheet(wb, "Corrections Applied", CORRECTIONS_HEADERS, [])
    return wb


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create an Evidence Scan Citation Ledger workbook.")
    parser.add_argument("--claims", type=Path, required=True, help="Path to claims JSON list")
    parser.add_argument(
        "--references",
        type=Path,
        help="Optional path to references JSON list. Omit when reports have weak or missing citations.",
    )
    parser.add_argument(
        "--inconsistencies",
        type=Path,
        help="Optional path to cross-report inconsistencies JSON list",
    )
    parser.add_argument("--topic", required=True, help="Evidence scan topic")
    parser.add_argument(
        "--verification-tool",
        default="Perplexity or equivalent source-checking tool",
        help="Planned verification tool",
    )
    parser.add_argument("--notes", default="", help="Optional README notes")
    parser.add_argument("--output", type=Path, required=True, help="Output .xlsx path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    claims = load_json(args.claims)
    references = load_json(args.references)
    inconsistencies = load_json(args.inconsistencies)

    wb = create_workbook(
        topic=args.topic,
        claims=claims,
        references=references,
        inconsistencies=inconsistencies,
        verification_tool=args.verification_tool,
        notes=args.notes,
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Created initial Citation Ledger: {args.output}")
    print("Next validation: python scripts/validate_ledger.py \"{}\" --stage initial".format(args.output))


if __name__ == "__main__":
    main()
