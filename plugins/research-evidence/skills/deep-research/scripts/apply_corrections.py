#!/usr/bin/env python3
"""Apply verified-ledger corrections to a source research report.

Implements the two-pass workflow from references/correction-workflow.md:

  Pass 1 — Inline targeted substitution
    For every [P] / [X] row in the verified ledger that targets this report,
    locate the matching anchor in the source text and replace it with the
    Correction column contents. The matcher tries strategies in order:
      a. Verbatim substring (exact match against the claim text).
      b. Normalized substring — strip markdown emphasis (*, **, _, __, `),
         collapse whitespace, lowercase. Many extracted claims are verbatim
         except for italics / case / whitespace; this pass anchors those.
      c. Sentence-level fuzzy — split source into sentence-like chunks
         (markdown-line-aware, then sentence-bounded inside each line),
         compare normalized claim to each chunk and to N-sentence sliding
         windows. Pick the best chunk above --sentence-cutoff. This pass
         anchors paraphrased / sentence-rewritten claims, which is the
         realistic shape of [P] corrections from external verifiers.
      d. Paragraph-level fuzzy difflib — last resort, the original
         coarse-grained matcher. Kept for back-compat.

  Pass 2 — Per-document Corrections appendix
    For every [P] / [X] flag the inline pass could not safely anchor in this
    report's text, append a structured "Corrections appendix" entry at the
    end of the same corrected source report. The appendix is part of the
    corrected-source-report file; it is not optional.

Hard rule (encoded by the script's exit behavior):
  Every [P] / [X] flag for the target report must end up either inline (Pass 1)
  or in the Corrections appendix (Pass 2). A flag that lives only in the
  cross-document Corrections Applied changelog does not satisfy this rule.
  The script writes the appendix unconditionally for any flag the inline pass
  did not place.

Usage:
  python scripts/apply_corrections.py \\
    --ledger "[Topic] — Citation Ledger.xlsx" \\
    --report-id R1 \\
    --source "[Topic] — R1.md" \\
    --output "[Topic] — R1 (Corrected).md"

Inputs:
  --ledger           Path to the verified Citation Ledger .xlsx (Claims tab).
  --report-id        Report identifier to filter the Claims tab on (e.g. R1).
  --source           Path to the original source research report (.md or .txt).
  --output           Path to write the corrected source report.
  --overwrite        Overwrite --output if it exists. Default False.
  --fuzzy-cutoff     difflib SequenceMatcher ratio threshold for paragraph
                     fuzzy fallback. Default 0.80.
  --sentence-cutoff  difflib ratio threshold for the sentence-level matcher.
                     Default 0.55. Lower = more inline anchoring but more
                     risk of mis-anchoring. Tune empirically.
  --max-window       Maximum number of consecutive sentences to consider as
                     a single anchor (for multi-sentence claim text).
                     Default 3.

Behavior:
  - Reads the Claims tab of the ledger.
  - Filters rows whose Report column matches --report-id.
  - For rows with Verified in {[P], [X]}: tries Pass 1 strategies in order.
  - Each substitution is recorded with (start, end, replacement) coordinates
    on the full source text so successive corrections do not overlap. If a
    candidate anchor would overlap an already-placed substitution, the row
    falls through to the appendix.
  - Writes the corrected source report to --output:
      original prose with substitutions applied
      + a "## Corrections appendix" section at the end with one row per
        unanchored flag.

Notes:
  - Substitution preserves the original prose surrounding the matched anchor.
    The script does NOT regenerate or reflow paragraphs.
  - The script does not modify --source. Output goes to --output only.
  - Run scripts/validate_ledger.py --stage final after generating the
    corrected report and the final ledger to confirm the audit trail is
    intact end-to-end.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable, Sequence

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment dependency
    sys.stderr.write(
        "openpyxl is required: pip install openpyxl --break-system-packages\n"
    )
    raise SystemExit(2) from exc


CORRECTED_CODES = {"[P]", "[X]"}
APPENDIX_HEADING = "## Corrections appendix"


@dataclass
class LedgerRow:
    claim_id: str
    section: str
    claim: str
    verified: str
    verification_notes: str
    correction: str
    source_cited: str


@dataclass
class Substitution:
    start: int
    end: int
    replacement: str
    claim_id: str
    strategy: str  # "verbatim" | "normalized" | "sentence" | "paragraph"


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_claims(ledger_path: Path, report_id: str) -> list[LedgerRow]:
    """Read the Claims tab and filter to rows for a given Report id."""
    wb = openpyxl.load_workbook(ledger_path, data_only=True)
    if "Claims" not in wb.sheetnames:
        sys.stderr.write(f"error: Claims tab not found in {ledger_path}\n")
        raise SystemExit(2)
    ws = wb["Claims"]

    headers = [_norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {
        "Report",
        "Claim ID",
        "Claim",
        "Verified",
        "Correction",
    }
    missing = required - set(headers)
    if missing:
        sys.stderr.write(
            f"error: Claims tab missing required columns: {sorted(missing)}\n"
        )
        raise SystemExit(2)

    idx = {h: i for i, h in enumerate(headers)}
    rows: list[LedgerRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        report = _norm(row[idx["Report"]])
        if report != report_id:
            continue
        rows.append(
            LedgerRow(
                claim_id=_norm(row[idx["Claim ID"]]),
                section=_norm(row[idx.get("Section", -1)]) if "Section" in idx else "",
                claim=_norm(row[idx["Claim"]]),
                verified=_norm(row[idx["Verified"]]),
                verification_notes=_norm(
                    row[idx.get("Verification Notes", -1)]
                )
                if "Verification Notes" in idx
                else "",
                correction=_norm(row[idx["Correction"]]),
                source_cited=_norm(row[idx.get("Source Cited", -1)])
                if "Source Cited" in idx
                else "",
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------

# Markdown emphasis / inline-code markers we strip during normalization.
# Brackets / parens / punctuation are kept so cite strings still anchor.
_MARKDOWN_MARK_RE = re.compile(r"[*_`]")
_WS_RE = re.compile(r"\s+")


def _normalize_text(s: str) -> tuple[str, list[int]]:
    """Return (normalized_text, original_index_for_each_char).

    - Strips markdown emphasis markers (* _ `).
    - Collapses runs of whitespace to a single space.
    - Lowercases.

    The returned ``idx_map`` is the same length as ``normalized_text``;
    each entry points to the *starting* position of that character in the
    original source, so callers can translate a normalized-space match span
    back to a span in the original text.
    """
    out: list[str] = []
    idx_map: list[int] = []
    in_ws = False
    for i, ch in enumerate(s):
        if ch in "*_`":
            continue
        if ch.isspace():
            if not in_ws:
                out.append(" ")
                idx_map.append(i)
                in_ws = True
            continue
        in_ws = False
        out.append(ch.lower())
        idx_map.append(i)
    return "".join(out), idx_map


def _normalize_simple(s: str) -> str:
    """Cheap normalization for similarity comparisons (no idx_map)."""
    s = _MARKDOWN_MARK_RE.sub("", s)
    s = _WS_RE.sub(" ", s)
    return s.strip().lower()


# ---------------------------------------------------------------------------
# Anchor finders (Pass 1 strategies)
# ---------------------------------------------------------------------------


def find_verbatim_span(claim: str, source: str) -> tuple[int, int] | None:
    """Pass 1a: verbatim substring."""
    if not claim:
        return None
    pos = source.find(claim)
    if pos < 0:
        return None
    return pos, pos + len(claim)


def find_normalized_span(claim: str, source: str) -> tuple[int, int] | None:
    """Pass 1b: normalized substring (strip markdown / case / whitespace)."""
    if not claim:
        return None
    norm_src, src_map = _normalize_text(source)
    norm_claim, _ = _normalize_text(claim)
    norm_claim = norm_claim.strip()
    if len(norm_claim) < 15:
        # Too short to anchor reliably.
        return None
    pos = norm_src.find(norm_claim)
    if pos < 0:
        return None
    end_pos = pos + len(norm_claim) - 1
    if end_pos >= len(src_map):
        return None
    return src_map[pos], src_map[end_pos] + 1


# Sentence-boundary regex: end-of-sentence punctuation followed by whitespace
# and a capital / opening paren / opening bracket. Conservative on purpose:
# it under-segments rather than over-segments, which keeps multi-clause
# bullet items as one anchor.
_SENT_BOUNDARY_RE = re.compile(r"(?<=[.!?])\s+(?=[A-Z(\[])")


def _split_sentences_with_pos(text: str) -> list[tuple[int, int]]:
    """Split text into sentence-like spans, preserving positions.

    Strategy: split on line breaks first (markdown lines often correspond
    to logical chunks — bullets, headings, paragraph fragments), then
    split each line on sentence-ending punctuation. Empty / whitespace-only
    spans are dropped.
    """
    spans: list[tuple[int, int]] = []
    line_start = 0
    n = len(text)
    for m in re.finditer(r"\n", text):
        line_end = m.start()
        if line_end > line_start:
            spans.append((line_start, line_end))
        line_start = line_end + 1
    if line_start < n:
        spans.append((line_start, n))

    sentences: list[tuple[int, int]] = []
    for s, e in spans:
        chunk = text[s:e]
        sub_start = 0
        for m in _SENT_BOUNDARY_RE.finditer(chunk):
            sub_end = m.start()
            if chunk[sub_start:sub_end].strip():
                sentences.append((s + sub_start, s + sub_end))
            sub_start = m.end()
        if chunk[sub_start:].strip():
            sentences.append((s + sub_start, e))
    return sentences


def find_sentence_span(
    claim: str,
    source: str,
    sentences: list[tuple[int, int]],
    cutoff: float,
    max_window: int,
) -> tuple[int, int, float] | None:
    """Pass 1c: sentence-level fuzzy match with sliding window.

    Returns (start, end, ratio) on the *original* source text or None.
    The matcher compares normalized-claim against normalized
    sentence-spans, including sliding windows of up to ``max_window``
    consecutive sentences for multi-sentence claims.
    """
    if not claim or not sentences:
        return None
    norm_claim = _normalize_simple(claim)
    if len(norm_claim) < 20:
        return None

    best: tuple[float, int, int] | None = None  # (ratio, start, end)

    # Single-sentence pass first.
    norm_sents: list[str] = []
    for s, e in sentences:
        norm_sents.append(_normalize_simple(source[s:e]))

    for i, (s, e) in enumerate(sentences):
        ns = norm_sents[i]
        if len(ns) < 10:
            continue
        # Length-ratio prefilter: very different lengths are unlikely matches.
        len_ratio = min(len(ns), len(norm_claim)) / max(len(ns), len(norm_claim))
        if len_ratio < cutoff * 0.6:
            continue
        sm = SequenceMatcher(a=ns, b=norm_claim, autojunk=False)
        if sm.quick_ratio() < cutoff:
            continue
        ratio = sm.ratio()
        if ratio >= cutoff and (best is None or ratio > best[0]):
            best = (ratio, s, e)

    # Sliding-window pass for multi-sentence claims.
    for w in range(2, max_window + 1):
        for i in range(0, len(sentences) - w + 1):
            s = sentences[i][0]
            e = sentences[i + w - 1][1]
            window_norm = _normalize_simple(source[s:e])
            if len(window_norm) < 10:
                continue
            len_ratio = min(len(window_norm), len(norm_claim)) / max(
                len(window_norm), len(norm_claim)
            )
            if len_ratio < cutoff * 0.6:
                continue
            sm = SequenceMatcher(a=window_norm, b=norm_claim, autojunk=False)
            if sm.quick_ratio() < cutoff:
                continue
            ratio = sm.ratio()
            if ratio >= cutoff and (best is None or ratio > best[0]):
                best = (ratio, s, e)

    if best is None:
        return None
    return best[1], best[2], best[0]


def find_fuzzy_paragraph_span(
    claim: str, source: str, paragraphs: Sequence[tuple[int, int]], cutoff: float
) -> tuple[int, int] | None:
    """Pass 1d: paragraph-level fuzzy difflib (back-compat last resort).

    Returns the original-text span of the longest matching block within
    the best-scoring paragraph above ``cutoff``, mirroring the behavior
    of the pre-F2.1 implementation.
    """
    if not claim:
        return None
    best_ratio = 0.0
    best_span: tuple[int, int] | None = None
    for s, e in paragraphs:
        para = source[s:e]
        sm = SequenceMatcher(a=para, b=claim, autojunk=False)
        ratio = sm.ratio()
        if ratio < cutoff or ratio <= best_ratio:
            continue
        match = sm.find_longest_match(0, len(para), 0, len(claim))
        if match.size == 0:
            continue
        if match.size < max(20, int(0.3 * len(claim))):
            continue
        best_ratio = ratio
        best_span = (s + match.a, s + match.a + match.size)
    return best_span


def _split_paragraphs_with_pos(text: str) -> list[tuple[int, int]]:
    """Return list of (start, end) for each paragraph (split on blank lines)."""
    spans: list[tuple[int, int]] = []
    cursor = 0
    for m in re.finditer(r"\n\s*\n", text):
        end = m.start()
        if end > cursor:
            spans.append((cursor, end))
        cursor = m.end()
    if cursor < len(text):
        spans.append((cursor, len(text)))
    return spans


# ---------------------------------------------------------------------------
# Substitution machinery
# ---------------------------------------------------------------------------


def _overlaps(start: int, end: int, subs: Iterable[Substitution]) -> bool:
    for s in subs:
        if not (end <= s.start or start >= s.end):
            return True
    return False


def apply_inline_pass(
    source_text: str,
    rows: Iterable[LedgerRow],
    fuzzy_cutoff: float,
    sentence_cutoff: float,
    max_window: int,
) -> tuple[str, list[LedgerRow], list[Substitution]]:
    """Apply Pass 1 substitutions and return (new_text, unanchored, subs).

    Tries strategies in order: verbatim → normalized → sentence-fuzzy →
    paragraph-fuzzy. First match wins; subsequent strategies only fire if
    earlier ones missed. Substitutions are tracked by (start, end) span on
    the original source text and applied in reverse order so positions
    remain valid.
    """
    sentences = _split_sentences_with_pos(source_text)
    paragraphs = _split_paragraphs_with_pos(source_text)

    subs: list[Substitution] = []
    unanchored: list[LedgerRow] = []

    for row in rows:
        if row.verified not in CORRECTED_CODES:
            continue
        if not row.correction:
            unanchored.append(row)
            continue

        # Strategy 1a: verbatim substring.
        span = find_verbatim_span(row.claim, source_text)
        if span and not _overlaps(*span, subs):
            subs.append(
                Substitution(span[0], span[1], row.correction, row.claim_id, "verbatim")
            )
            continue

        # Strategy 1b: normalized substring.
        span = find_normalized_span(row.claim, source_text)
        if span and not _overlaps(*span, subs):
            subs.append(
                Substitution(span[0], span[1], row.correction, row.claim_id, "normalized")
            )
            continue

        # Strategy 1c: sentence-level fuzzy.
        sent = find_sentence_span(
            row.claim, source_text, sentences, sentence_cutoff, max_window
        )
        if sent and not _overlaps(sent[0], sent[1], subs):
            subs.append(
                Substitution(sent[0], sent[1], row.correction, row.claim_id, "sentence")
            )
            continue

        # Strategy 1d: paragraph-level fuzzy difflib (last-resort).
        span = find_fuzzy_paragraph_span(row.claim, source_text, paragraphs, fuzzy_cutoff)
        if span and not _overlaps(*span, subs):
            subs.append(
                Substitution(span[0], span[1], row.correction, row.claim_id, "paragraph")
            )
            continue

        unanchored.append(row)

    # Apply substitutions in reverse position order to keep offsets valid.
    out = source_text
    for s in sorted(subs, key=lambda x: x.start, reverse=True):
        out = out[: s.start] + s.replacement + out[s.end :]
    return out, unanchored, subs


def render_appendix(unanchored: Sequence[LedgerRow]) -> str:
    """Render the per-document Corrections appendix block."""
    if not unanchored:
        return ""
    lines = [
        "",
        "---",
        "",
        APPENDIX_HEADING,
        "",
        (
            "The following corrections from the verified Citation Ledger apply "
            "to this report but could not be safely anchored to a specific "
            "passage in the original text. They are included here so that this "
            "corrected report remains a self-contained record of all "
            "corrections applied."
        ),
        "",
        "| Claim ID | Verified code | Location reference | Correction | Source |",
        "|---|---|---|---|---|",
    ]
    for row in unanchored:
        location = row.section or "(section not recorded — see verification notes)"
        notes = row.verification_notes or ""
        if notes and notes != location:
            location = f"{location} — {notes}"
        cell = lambda s: s.replace("|", "\\|").replace("\n", " ").strip()
        lines.append(
            "| "
            + " | ".join(
                [
                    cell(row.claim_id),
                    cell(row.verified),
                    cell(location),
                    cell(row.correction),
                    cell(row.source_cited or "(see ledger)"),
                ]
            )
            + " |"
        )
    lines.append("")
    return "\n".join(lines)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Apply verified-ledger corrections to a source report."
    )
    parser.add_argument("--ledger", required=True, type=Path)
    parser.add_argument("--report-id", required=True)
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite --output if it already exists. Default False.",
    )
    parser.add_argument(
        "--fuzzy-cutoff",
        type=float,
        default=0.80,
        help="Paragraph-fuzzy difflib cutoff (last resort). Default 0.80.",
    )
    parser.add_argument(
        "--sentence-cutoff",
        type=float,
        default=0.55,
        help="Sentence-fuzzy difflib cutoff. Default 0.55.",
    )
    parser.add_argument(
        "--max-window",
        type=int,
        default=3,
        help="Max consecutive sentences in the sliding-window matcher. Default 3.",
    )
    args = parser.parse_args(argv)

    if not args.ledger.exists():
        sys.stderr.write(f"error: ledger not found: {args.ledger}\n")
        return 2
    if not args.source.exists():
        sys.stderr.write(f"error: source not found: {args.source}\n")
        return 2
    if args.output.exists() and not args.overwrite:
        sys.stderr.write(
            f"error: --output already exists: {args.output}\n"
            "       Pass --overwrite to replace it, or choose a different path.\n"
        )
        return 2

    rows = read_claims(args.ledger, args.report_id)
    if not rows:
        sys.stderr.write(
            f"warning: no Claims rows for report {args.report_id} in {args.ledger}\n"
        )

    source_text = args.source.read_text(encoding="utf-8")

    corrected, unanchored, subs = apply_inline_pass(
        source_text,
        rows,
        args.fuzzy_cutoff,
        args.sentence_cutoff,
        args.max_window,
    )

    appendix = render_appendix(unanchored)
    if appendix:
        corrected = corrected.rstrip() + "\n" + appendix

    # If --overwrite is set and the existing target is read-only (e.g., copied
    # from a read-only upload), unlink it first so the new file gets fresh
    # permissions instead of failing with PermissionError on write.
    if args.overwrite and args.output.exists():
        try:
            args.output.unlink()
        except OSError as e:
            sys.stderr.write(
                f"error: could not remove existing --output before overwrite: {e}\n"
            )
            return 2

    args.output.write_text(corrected, encoding="utf-8")

    total_corrected_rows = sum(1 for r in rows if r.verified in CORRECTED_CODES)
    inline_count = total_corrected_rows - len(unanchored)
    by_strategy: dict[str, int] = {}
    for s in subs:
        by_strategy[s.strategy] = by_strategy.get(s.strategy, 0) + 1
    breakdown = ", ".join(f"{k}={v}" for k, v in sorted(by_strategy.items())) or "(none)"
    sys.stdout.write(
        f"applied {inline_count}/{total_corrected_rows} corrections inline "
        f"[{breakdown}]; "
        f"{len(unanchored)} written to Corrections appendix in {args.output}\n"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
