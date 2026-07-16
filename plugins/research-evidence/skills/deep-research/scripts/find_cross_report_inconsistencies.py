#!/usr/bin/env python3
"""Mechanically surface cross-report inconsistency candidates from a Citation Ledger.

Implements the "Mechanical filter recipe" from references/phase-2-verify-prep.md:

  1. Normalize each row's Source Cited string to a set of (author, year)
     tuples and skip sentinel values (Author synthesis, MISSING_IN_REPORT,
     uncited). Two rows belong to the same group if their tuple sets
     intersect — so "MDRC, 2018, 2020, 2023" and "MDRC, 2020, 2023" group
     together via the shared (mdrc, 2020) tuple.
  2. For groups with rows from >=2 reports, scan for divergent numeric
     tokens, divergent dates / time periods, divergent claim_type values,
     divergent Evidence Type values, and opposite-direction effect language.
  3. Emit one candidate row per (group, divergence pattern) for the human
     or verifier to triage into the Cross-Report Inconsistencies tab.

The script is a *filter*, not a judgment. It surfaces candidates; a human
or verifier decides which candidates land in the ledger as real
inconsistencies. The output is a CSV (or stdout table) so the user can
paste it into the Cross-Report Inconsistencies tab after triage.

Usage:
  python scripts/find_cross_report_inconsistencies.py \\
    --ledger "[Topic] — Citation Ledger.xlsx" \\
    --output "[Topic] — Inconsistency Candidates.csv"

  # or, write to stdout instead of a CSV:
  python scripts/find_cross_report_inconsistencies.py \\
    --ledger "[Topic] — Citation Ledger.xlsx"

Inputs:
  --ledger        Path to the Citation Ledger .xlsx (initial or verified).
  --output        Optional path to write a CSV of candidates. If omitted,
                  candidates print to stdout as a Markdown table.
  --min-reports   Minimum number of distinct reports per Source Cited group
                  to be considered for cross-report comparison. Default 2.

Detected divergence patterns (mirrors the conflict-pattern checklist):
  - same-source-different-magnitude: distinct numeric tokens in Claim
  - same-source-different-time-point: distinct date / year / time-period
    tokens in Claim or evidence_detail
  - same-source-different-claim-domain: distinct claim_type values
  - composite-vs-primary-divergence: distinct Evidence Type values
  - direction-of-effect inversion: opposite-sign effect language
  - cross-report-citation-reuse: same source cited by >=2 reports for any
    claim, even when no other divergence is detected (verifier should
    confirm the source supports all uses)

This script does not detect:
  - adjacent-domain leak (requires semantic comparison, not in scope)
  - internal numeric variance within a single report (different filter shape;
    the same recipe can be applied with min-reports=1 if useful, but the
    default focus here is cross-report)
  - working-paper-vs-publication year drift (same author, different year —
    e.g., "Weiss, Bloom, & Singh, 2022" working-paper version vs "(2023)"
    EEPA publication). The (author, year) tuple normalization treats these
    as distinct sources by design; catching them requires a year-tolerance
    layer the verifier handles directly. Run 3 X06 is the known case.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Sequence

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment dependency
    sys.stderr.write(
        "openpyxl is required: pip install openpyxl --break-system-packages\n"
    )
    raise SystemExit(2) from exc


NUMERIC_TOKEN_RE = re.compile(
    r"""
    (?<![A-Za-z])              # not preceded by a letter
    (?:                        # number forms:
        \d+\.\d+               # decimal:        12.5
        |\d+(?:,\d{3})+        # comma-grouped:  12,000
        |\d+%                  # percentage:     30%
        |\$\d+(?:[.,]\d+)?     # dollar:         $30, $30.5
        |\d+                   # integer:        30
    )
    """,
    re.VERBOSE,
)

# Years and obvious time-period tokens.
TIME_TOKEN_RE = re.compile(
    r"""
    \b(?:
        (?:19|20)\d{2}                                          # year
        |\d+\s*(?:year|yr|month|week|day)s?\b                   # duration
        |(?:6|8|10|12|24|36|48)-year\b                          # follow-up
        |(?:short|medium|long)-term\b
        |baseline|endpoint|follow[- ]up
    )
    """,
    re.VERBOSE | re.IGNORECASE,
)

POSITIVE_EFFECT_RE = re.compile(
    r"\b(?:increase[ds]?|improv(?:e[ds]?|ing|ement)|gain[s]?|"
    r"rise[ns]?|grow[s]?|positive\s+effect|raise[ds]?|boost[s]?)\b",
    re.IGNORECASE,
)
NEGATIVE_EFFECT_RE = re.compile(
    r"\b(?:decrease[ds]?|declin(?:e[ds]?|ing)|reduc(?:e[ds]?|tion)|"
    r"drop[s]?|fall[s]?|fell|negative\s+effect|worsen(?:ed|ing)?|"
    r"lower(?:ed|ing)?|harm[s]?)\b",
    re.IGNORECASE,
)


@dataclass
class ClaimRow:
    report: str
    claim_id: str
    section: str
    claim: str
    source_cited: str
    evidence_type: str
    claim_type: str
    evidence_detail: str


# Sentinel `Source Cited` values that should NOT be grouped — author-synthesis
# rows, intentionally-blank rows, and missing-in-report flags. These create
# noise candidates if grouped on the literal string match (Run 3 X03/X04/X05).
_SENTINEL_CITES = {
    "",
    "n/a",
    "na",
    "—",
    "-",
    "uncited",
    "author synthesis",
    "author synthesis, no citation",
    "missing_in_report",
    "missing from report",
    "missing",
    "(see ledger)",
    "see ledger",
}

# Common particles / role-words to skip when extracting first author surname.
_SURNAME_SKIP = {
    "the",
    "and",
    "et",
    "al",
    "ed",
    "eds",
    "von",
    "van",
    "de",
    "la",
    "le",
    "du",
    "del",
    "da",
}

_YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")


def normalize_cite(raw: str) -> set[tuple[str, str]]:
    """Return a set of (author_surname_lower, year) tuples for a cite string.

    Handles common author-cite-string variation across reports:
      "Bird et al., 2021"                 → {(bird, 2021)}
      "Bird et al. (2021, JEBO)"          → {(bird, 2021)}
      "Castleman & Page, 2014, 2016"      → {(castleman, 2014), (castleman, 2016)}
      "MDRC, 2018, 2020, 2023, 2025"      → {(mdrc, 2018), (mdrc, 2020), ...}
      "Weiss, Bloom, & Singh (2023)"      → {(weiss, 2023)}

    Returns an empty set for sentinel values (Author synthesis, uncited,
    MISSING_IN_REPORT, empty / dash) and for cites with no extractable
    year. Empty sets are excluded from grouping.
    """
    if raw is None:
        return set()
    s = str(raw).strip()
    if not s:
        return set()
    if s.lower().strip(",.;:") in _SENTINEL_CITES:
        return set()

    years = _YEAR_RE.findall(s)
    if not years:
        return set()

    # Take the head of the cite (before first comma / & / "and" / "et al").
    # Then take the first non-particle alphabetic token as the surname.
    head = re.split(r",|&|\bet al\b|\band\b", s, maxsplit=1)[0]
    # Strip leading punctuation, parens, quotes.
    head = re.sub(r"^[^A-Za-z]+", "", head)
    tokens = [
        re.sub(r"[^A-Za-z]", "", t).lower()
        for t in re.split(r"\s+", head)
        if t.strip()
    ]
    surname = ""
    for tok in tokens:
        if not tok or tok in _SURNAME_SKIP:
            continue
        surname = tok
        break
    if not surname:
        return set()

    return {(surname, y) for y in set(years)}


@dataclass
class Candidate:
    inconsistency_label: str
    pattern: str
    source_cited: str
    contributing_reports: list[str] = field(default_factory=list)
    contributing_claim_ids: list[str] = field(default_factory=list)
    note: str = ""


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_claims(ledger_path: Path) -> list[ClaimRow]:
    wb = openpyxl.load_workbook(ledger_path, data_only=True)
    if "Claims" not in wb.sheetnames:
        sys.stderr.write(f"error: Claims tab not found in {ledger_path}\n")
        raise SystemExit(2)
    ws = wb["Claims"]

    headers = [_norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {"Report", "Claim ID", "Claim", "Source Cited", "Evidence Type"}
    missing = required - set(headers)
    if missing:
        sys.stderr.write(
            f"error: Claims tab missing required columns: {sorted(missing)}\n"
        )
        raise SystemExit(2)

    idx = {h: i for i, h in enumerate(headers)}

    def cell(row, col_name: str) -> str:
        i = idx.get(col_name)
        if i is None:
            return ""
        return _norm(row[i])

    rows: list[ClaimRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        report = cell(row, "Report")
        claim_id = cell(row, "Claim ID")
        if not report or not claim_id:
            continue
        rows.append(
            ClaimRow(
                report=report,
                claim_id=claim_id,
                section=cell(row, "Section"),
                claim=cell(row, "Claim"),
                source_cited=cell(row, "Source Cited"),
                evidence_type=cell(row, "Evidence Type"),
                claim_type=cell(row, "claim_type"),  # enrichment field if preserved
                evidence_detail=cell(row, "evidence_detail"),
            )
        )
    return rows


def _direction(text: str) -> str | None:
    pos = bool(POSITIVE_EFFECT_RE.search(text))
    neg = bool(NEGATIVE_EFFECT_RE.search(text))
    if pos and not neg:
        return "positive"
    if neg and not pos:
        return "negative"
    return None


def _build_groups(rows: list[ClaimRow]) -> list[tuple[str, list[ClaimRow]]]:
    """Group rows by overlapping normalized cite tuples.

    Two rows belong to the same group if their normalized cite sets share
    at least one (author, year) tuple. This catches author-cite-string
    variation that would defeat literal-string grouping (e.g.,
    "MDRC, 2018, 2020, 2023" and "MDRC, 2020, 2023" share (mdrc, 2020)).

    Sentinel cites (Author synthesis, uncited, MISSING_IN_REPORT, blank)
    are dropped before grouping — they are not real source attributions
    and grouping on them creates noise candidates.

    Returns a list of (representative_cite_string, rows) pairs. The
    representative is the most-common Source Cited string in the group.
    """
    indexed: list[tuple[int, ClaimRow, set[tuple[str, str]]]] = []
    for i, r in enumerate(rows):
        cite_set = normalize_cite(r.source_cited)
        if cite_set:
            indexed.append((i, r, cite_set))

    parent: dict[int, int] = {i: i for i, _, _ in indexed}

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    # Inverted index: tuple → list of row indices that contain it.
    by_tuple: dict[tuple[str, str], list[int]] = defaultdict(list)
    for i, _, cs in indexed:
        for t in cs:
            by_tuple[t].append(i)
    for ids in by_tuple.values():
        for j in range(1, len(ids)):
            union(ids[0], ids[j])

    groups_by_root: dict[int, list[ClaimRow]] = defaultdict(list)
    for i, r, _ in indexed:
        groups_by_root[find(i)].append(r)

    out: list[tuple[str, list[ClaimRow]]] = []
    for group in groups_by_root.values():
        # Representative cite string: most common literal value in the group.
        cite_counts: dict[str, int] = defaultdict(int)
        for r in group:
            cite_counts[r.source_cited] += 1
        rep = max(cite_counts.items(), key=lambda kv: kv[1])[0]
        out.append((rep, group))
    return out


def detect_candidates(rows: list[ClaimRow], min_reports: int) -> list[Candidate]:
    candidates: list[Candidate] = []
    grouped = _build_groups(rows)

    counter = 0

    def next_label() -> str:
        nonlocal counter
        counter += 1
        return f"X{counter:02d}"

    for source, group in grouped:
        reports_in_group = sorted({r.report for r in group})
        if len(reports_in_group) < min_reports:
            continue

        # Always emit a citation-reuse candidate so verifier can confirm
        # the source supports all uses across reports.
        candidates.append(
            Candidate(
                inconsistency_label=next_label(),
                pattern="cross-report-citation-reuse",
                source_cited=source,
                contributing_reports=reports_in_group,
                contributing_claim_ids=sorted({r.claim_id for r in group}),
                note=(
                    "Same source cited by >=2 reports — verifier should "
                    "confirm the source supports all uses, even if no "
                    "numeric / domain divergence is detected."
                ),
            )
        )

        # Numeric magnitude divergence.
        per_report_numbers: dict[str, set[str]] = defaultdict(set)
        for r in group:
            for token in NUMERIC_TOKEN_RE.findall(r.claim):
                per_report_numbers[r.report].add(token)
        all_numbers = {n for nums in per_report_numbers.values() for n in nums}
        if all_numbers and len({frozenset(v) for v in per_report_numbers.values()}) > 1:
            candidates.append(
                Candidate(
                    inconsistency_label=next_label(),
                    pattern="same-source-different-magnitude",
                    source_cited=source,
                    contributing_reports=reports_in_group,
                    contributing_claim_ids=sorted({r.claim_id for r in group}),
                    note=(
                        "Distinct numeric tokens across reports for the same "
                        "source: "
                        + "; ".join(
                            f"{rep}: {sorted(nums)}"
                            for rep, nums in per_report_numbers.items()
                        )
                    ),
                )
            )

        # Time-point divergence.
        per_report_times: dict[str, set[str]] = defaultdict(set)
        for r in group:
            blob = f"{r.claim} {r.evidence_detail}"
            for token in TIME_TOKEN_RE.findall(blob):
                per_report_times[r.report].add(token.lower())
        if (
            per_report_times
            and len({frozenset(v) for v in per_report_times.values()}) > 1
        ):
            candidates.append(
                Candidate(
                    inconsistency_label=next_label(),
                    pattern="same-source-different-time-point",
                    source_cited=source,
                    contributing_reports=reports_in_group,
                    contributing_claim_ids=sorted({r.claim_id for r in group}),
                    note=(
                        "Distinct time-period / date tokens across reports: "
                        + "; ".join(
                            f"{rep}: {sorted(toks)}"
                            for rep, toks in per_report_times.items()
                        )
                    ),
                )
            )

        # Claim-domain divergence.
        per_report_claim_types: dict[str, set[str]] = defaultdict(set)
        for r in group:
            if r.claim_type:
                per_report_claim_types[r.report].add(r.claim_type)
        if (
            per_report_claim_types
            and len({frozenset(v) for v in per_report_claim_types.values()}) > 1
        ):
            candidates.append(
                Candidate(
                    inconsistency_label=next_label(),
                    pattern="same-source-different-claim-domain",
                    source_cited=source,
                    contributing_reports=reports_in_group,
                    contributing_claim_ids=sorted({r.claim_id for r in group}),
                    note=(
                        "Distinct claim_type values across reports — same "
                        "source supporting different claim domains: "
                        + "; ".join(
                            f"{rep}: {sorted(v)}"
                            for rep, v in per_report_claim_types.items()
                        )
                    ),
                )
            )

        # Composite-vs-primary divergence (Evidence Type mismatch).
        per_report_etypes: dict[str, set[str]] = defaultdict(set)
        for r in group:
            if r.evidence_type:
                per_report_etypes[r.report].add(r.evidence_type)
        if (
            per_report_etypes
            and len({frozenset(v) for v in per_report_etypes.values()}) > 1
        ):
            candidates.append(
                Candidate(
                    inconsistency_label=next_label(),
                    pattern="composite-vs-primary-divergence",
                    source_cited=source,
                    contributing_reports=reports_in_group,
                    contributing_claim_ids=sorted({r.claim_id for r in group}),
                    note=(
                        "Distinct Evidence Type values across reports for "
                        "the same source: "
                        + "; ".join(
                            f"{rep}: {sorted(v)}"
                            for rep, v in per_report_etypes.items()
                        )
                    ),
                )
            )

        # Direction-of-effect inversion.
        per_report_dirs: dict[str, set[str]] = defaultdict(set)
        for r in group:
            d = _direction(r.claim)
            if d:
                per_report_dirs[r.report].add(d)
        directions = {d for vs in per_report_dirs.values() for d in vs}
        if "positive" in directions and "negative" in directions:
            candidates.append(
                Candidate(
                    inconsistency_label=next_label(),
                    pattern="direction-of-effect-inversion",
                    source_cited=source,
                    contributing_reports=reports_in_group,
                    contributing_claim_ids=sorted({r.claim_id for r in group}),
                    note=(
                        "Opposite-direction effect language across reports: "
                        + "; ".join(
                            f"{rep}: {sorted(v)}"
                            for rep, v in per_report_dirs.items()
                        )
                    ),
                )
            )
    return candidates


def write_csv(candidates: Sequence[Candidate], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "Inconsistency ID (candidate)",
                "Pattern",
                "Source Cited",
                "Contributing reports",
                "Contributing Claim IDs",
                "Note",
            ]
        )
        for c in candidates:
            writer.writerow(
                [
                    c.inconsistency_label,
                    c.pattern,
                    c.source_cited,
                    ", ".join(c.contributing_reports),
                    ", ".join(c.contributing_claim_ids),
                    c.note,
                ]
            )


def write_markdown(candidates: Sequence[Candidate]) -> str:
    if not candidates:
        return "_No candidate inconsistencies detected._\n"
    lines = [
        "| Inconsistency ID (candidate) | Pattern | Source Cited | Reports | Claim IDs | Note |",
        "|---|---|---|---|---|---|",
    ]
    for c in candidates:
        cell = lambda s: str(s).replace("|", "\\|").replace("\n", " ").strip()
        lines.append(
            "| "
            + " | ".join(
                [
                    cell(c.inconsistency_label),
                    cell(c.pattern),
                    cell(c.source_cited),
                    cell(", ".join(c.contributing_reports)),
                    cell(", ".join(c.contributing_claim_ids)),
                    cell(c.note),
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Surface cross-report inconsistency candidates from a Citation Ledger."
    )
    parser.add_argument("--ledger", required=True, type=Path)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--min-reports", type=int, default=2)
    args = parser.parse_args(argv)

    if not args.ledger.exists():
        sys.stderr.write(f"error: ledger not found: {args.ledger}\n")
        return 2

    rows = read_claims(args.ledger)
    candidates = detect_candidates(rows, args.min_reports)

    if args.output is not None:
        write_csv(candidates, args.output)
        sys.stdout.write(
            f"wrote {len(candidates)} candidate(s) to {args.output}\n"
        )
    else:
        sys.stdout.write(write_markdown(candidates))
        sys.stdout.write(
            f"\n_{len(candidates)} candidate(s) detected._\n"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
