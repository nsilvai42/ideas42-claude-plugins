#!/usr/bin/env python3
"""Validate an Evidence Scan Citation Ledger workbook.

Use this helper after Phase 2 ledger creation, after external verification,
and after Phase 3 correction application. It checks workbook structure,
required columns, claim ID uniqueness, report labels, verification codes,
stakes values, and correction completeness.

Validation stages:
    initial  - Phase 2 ledger before verification. Verification fields should be blank.
    verified - Ledger returned from external verification. Every row should have a verification code; non-[C] rows need notes; [P] and [X] rows need corrections.
    final    - Phase 3 final ledger. Corrected rows need corrections, Corrections Applied rows, and Claude QA.

Examples:
    python scripts/validate_ledger.py "Benefits chatbot evidence scan — Citation Ledger.xlsx" --stage initial
    python scripts/validate_ledger.py "Benefits chatbot evidence scan — Citation Ledger.xlsx" --stage verified
    python scripts/validate_ledger.py "Benefits chatbot evidence scan — Citation Ledger (Final).xlsx" --stage final
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

REQUIRED_TABS = [
    "README",
    "Claims",
    "References",
    "Cross-Report Inconsistencies",
    "Corrections Applied",
]

REQUIRED_COLUMNS = {
    "Claims": [
        "Report",
        "Claim ID",
        "Section",
        "Claim",
        "Source Cited",
        "Full Citation",
        "Evidence Type",
        "Stakes",
        "Verified",
        "Verification Notes",
        "Correction",
        "Claude QA",
    ],
    "References": ["Report", "Reference ID", "Full Citation", "Referenced in claims"],
    "Cross-Report Inconsistencies": [
        "Inconsistency ID",
        "Source scope",
        "Report",
        "Claim text",
        "Affected Claim IDs",
        "Type",
        "Resolution",
    ],
    "Corrections Applied": [
        "Claim ID",
        "Verified code",
        "Original text",
        "Corrected text",
        "Reason",
        "Claude QA",
    ],
}

VALID_STAKES = {"H", "M", "L"}
VALID_VERIFIED = {"", "[C]", "[P]", "[U]", "[X]"}
CORRECTION_CODES = {"[P]", "[X]"}
NON_CONFIRMED_CODES = {"[P]", "[U]", "[X]"}
RECOMMENDED_REPORT_LABELS = {"R1", "R2", "R3", "R4", "R5"}
INITIAL_BLANK_FIELDS = ["Verified", "Verification Notes", "Correction", "Claude QA"]


def cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def get_headers(ws) -> list[str]:
    return [cell_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def row_dicts(ws) -> Iterable[dict[str, str]]:
    headers = get_headers(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell_text(value) for value in row):
            continue
        yield {header: cell_text(value) for header, value in zip(headers, row)}


def require_tabs(wb) -> list[str]:
    errors: list[str] = []
    for tab in REQUIRED_TABS:
        if tab not in wb.sheetnames:
            errors.append(f"Missing required tab: {tab}")
    return errors


def require_columns(wb) -> list[str]:
    errors: list[str] = []
    for tab, required in REQUIRED_COLUMNS.items():
        if tab not in wb.sheetnames:
            continue
        headers = get_headers(wb[tab])
        missing = [column for column in required if column not in headers]
        if missing:
            errors.append(f"{tab}: missing required columns: {', '.join(missing)}")
    return errors


def validate_report_labels(rows: list[dict[str, str]]) -> list[str]:
    warnings: list[str] = []
    reports = sorted({row.get("Report", "") for row in rows if row.get("Report", "")})
    extra = [report for report in reports if report not in RECOMMENDED_REPORT_LABELS]
    if extra or len(reports) > 5:
        warnings.append(
            "Ledger includes report labels beyond the usual R1 through R5 workflow. "
            "Batch reports or prioritize the highest-value five if verification becomes unwieldy. "
            f"Reports found: {', '.join(reports)}"
        )
    return warnings


def validate_initial_stage(index: int, row: dict[str, str]) -> list[str]:
    errors: list[str] = []
    for field in INITIAL_BLANK_FIELDS:
        if row.get(field, ""):
            errors.append(f"Claims row {index}: initial ledger should leave {field} blank")
    return errors


def validate_verified_or_final_stage(index: int, row: dict[str, str], stage: str) -> list[str]:
    errors: list[str] = []
    verified = row.get("Verified", "")
    notes = row.get("Verification Notes", "")
    correction = row.get("Correction", "")
    qa = row.get("Claude QA", "")

    if not verified:
        errors.append(f"Claims row {index}: {stage} ledger should have Verified filled")
        return errors

    if verified in NON_CONFIRMED_CODES and not notes:
        errors.append(f"Claims row {index}: {verified} row missing Verification Notes")

    if verified in CORRECTION_CODES and not correction:
        errors.append(f"Claims row {index}: {verified} row missing Correction")

    if stage == "final" and verified in CORRECTION_CODES and not qa:
        errors.append(f"Claims row {index}: {verified} row missing Claude QA")

    return errors


def validate_claims(wb, stage: str) -> tuple[list[str], list[str]]:
    if "Claims" not in wb.sheetnames:
        return [], []

    errors: list[str] = []
    seen_claim_ids: set[str] = set()
    rows = list(row_dicts(wb["Claims"]))

    for index, row in enumerate(rows, start=2):
        claim_id = row.get("Claim ID", "")
        report = row.get("Report", "")
        claim = row.get("Claim", "")
        evidence_type = row.get("Evidence Type", "")
        stakes = row.get("Stakes", "")
        verified = row.get("Verified", "")

        for field, value in [
            ("Report", report),
            ("Claim ID", claim_id),
            ("Claim", claim),
            ("Evidence Type", evidence_type),
            ("Stakes", stakes),
        ]:
            if not value:
                errors.append(f"Claims row {index}: missing {field}")

        if claim_id:
            if claim_id in seen_claim_ids:
                errors.append(f"Claims row {index}: duplicate Claim ID {claim_id}")
            seen_claim_ids.add(claim_id)

        if stakes and stakes not in VALID_STAKES:
            errors.append(f"Claims row {index}: invalid Stakes value {stakes}; use H, M, or L")

        if verified and verified not in VALID_VERIFIED:
            errors.append(
                f"Claims row {index}: invalid Verified value {verified}; use [C], [P], [U], [X], or blank"
            )

        if stage == "initial":
            errors.extend(validate_initial_stage(index, row))
        elif stage in {"verified", "final"}:
            errors.extend(validate_verified_or_final_stage(index, row, stage))

    warnings = validate_report_labels(rows)
    return errors, warnings


def validate_references(wb) -> list[str]:
    if "References" not in wb.sheetnames:
        return []

    errors: list[str] = []
    seen_reference_ids: set[tuple[str, str]] = set()

    for index, row in enumerate(row_dicts(wb["References"]), start=2):
        report = row.get("Report", "")
        reference_id = row.get("Reference ID", "")
        full_citation = row.get("Full Citation", "")

        if not report:
            errors.append(f"References row {index}: missing Report")
        if not reference_id:
            errors.append(f"References row {index}: missing Reference ID")
        if not full_citation:
            errors.append(f"References row {index}: missing Full Citation")

        if report and reference_id:
            key = (report, reference_id)
            if key in seen_reference_ids:
                errors.append(f"References row {index}: duplicate Reference ID {reference_id} for {report}")
            seen_reference_ids.add(key)

    return errors


def validate_corrections_applied(wb, stage: str) -> list[str]:
    if stage != "final":
        return []

    if "Claims" not in wb.sheetnames or "Corrections Applied" not in wb.sheetnames:
        return []

    errors: list[str] = []
    corrected_claim_ids = {
        row.get("Claim ID", "")
        for row in row_dicts(wb["Claims"])
        if row.get("Verified", "") in CORRECTION_CODES
    }
    all_claim_ids = {
        row.get("Claim ID", "")
        for row in row_dicts(wb["Claims"])
        if row.get("Claim ID", "")
    }
    corrections_rows = list(row_dicts(wb["Corrections Applied"]))
    corrections_tab_ids = {
        row.get("Claim ID", "")
        for row in corrections_rows
        if row.get("Claim ID", "")
    }

    missing = sorted(corrected_claim_ids - corrections_tab_ids)
    if missing:
        errors.append(
            "Corrections Applied tab missing corrected Claim IDs: " + ", ".join(missing)
        )

    extra = sorted(corrections_tab_ids - corrected_claim_ids)
    if extra:
        errors.append(
            "Corrections Applied tab includes Claim IDs that are not [P] or [X] in Claims: "
            + ", ".join(extra)
        )

    unknown = sorted(corrections_tab_ids - all_claim_ids)
    if unknown:
        errors.append(
            "Corrections Applied tab includes Claim IDs not found in Claims: " + ", ".join(unknown)
        )

    for index, row in enumerate(corrections_rows, start=2):
        claim_id = row.get("Claim ID", "")
        verified_code = row.get("Verified code", "")
        original_text = row.get("Original text", "")
        corrected_text = row.get("Corrected text", "")
        reason = row.get("Reason", "")
        qa = row.get("Claude QA", "")

        if not claim_id:
            errors.append(f"Corrections Applied row {index}: missing Claim ID")
        if verified_code not in CORRECTION_CODES:
            errors.append(
                f"Corrections Applied row {index}: invalid Verified code {verified_code}; use [P] or [X]"
            )
        if not original_text:
            errors.append(f"Corrections Applied row {index}: missing Original text")
        if not corrected_text:
            errors.append(f"Corrections Applied row {index}: missing Corrected text")
        if not reason:
            errors.append(f"Corrections Applied row {index}: missing Reason")
        if not qa:
            errors.append(f"Corrections Applied row {index}: missing Claude QA")

    return errors


def validate_workbook(path: Path, stage: str) -> tuple[list[str], list[str]]:
    wb = load_workbook(path)
    errors: list[str] = []
    warnings: list[str] = []

    errors.extend(require_tabs(wb))
    errors.extend(require_columns(wb))
    claim_errors, claim_warnings = validate_claims(wb, stage)
    errors.extend(claim_errors)
    warnings.extend(claim_warnings)
    errors.extend(validate_references(wb))
    errors.extend(validate_corrections_applied(wb, stage))
    return errors, warnings


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate an Evidence Scan Citation Ledger workbook.")
    parser.add_argument("ledger", type=Path, help="Path to Citation Ledger .xlsx")
    parser.add_argument(
        "--stage",
        choices=["initial", "verified", "final"],
        default="final",
        help="Validation strictness: initial, verified, or final. Defaults to final.",
    )
    args = parser.parse_args()

    if not args.ledger.exists():
        raise SystemExit(f"File not found: {args.ledger}")

    errors, warnings = validate_workbook(args.ledger, args.stage)

    if warnings:
        print("Citation Ledger validation warnings:")
        for warning in warnings:
            print(f"- {warning}")

    if errors:
        print("Citation Ledger validation failed:")
        for error in errors:
            print(f"- {error}")
        raise SystemExit(1)

    print(f"Citation Ledger validation passed for stage '{args.stage}': {args.ledger}")


if __name__ == "__main__":
    main()
